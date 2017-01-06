#! python3

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 42
sheet['B1'] = 84

sheet2 = wb.create_sheet() #defaults are Sheet, then Sheet1
sheet2.title = 'Sheet 2'
#can also create sheets at index/with title
sheet0 = wb.create_sheet(index=0,title="The sheet before all others")
#recommended that you always save a new copy, not overwrite original
wb.save('workingBook.xlsx')